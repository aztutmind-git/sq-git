import io
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session

import models
import schemas
from database import get_db
from deps import get_current_user, require_admin

router = APIRouter(prefix="/api", tags=["questions"])

VALID_SUBJECTS = {"chemistry", "physics", "botany", "zoology", "commerce", "accounts", "mathematics", "nutrition"}
LETTER_MAP = {"A": 0, "B": 1, "C": 2, "D": 3, "1": 0, "2": 1, "3": 2, "4": 3}


# ---------- student-facing: fetch a shuffled set of questions for a level ----------
@router.get("/questions", response_model=List[schemas.QuestionForQuiz])
def get_questions_for_quiz(
    subject: str, level: int,
    db: Session = Depends(get_db), _user: models.User = Depends(get_current_user),
):
    if subject not in VALID_SUBJECTS:
        raise HTTPException(status_code=400, detail="Unknown subject")
    rows = db.query(models.Question).filter(
        models.Question.subject == subject, models.Question.level == level
    ).all()
    return [
        schemas.QuestionForQuiz(
            id=q.id, subject=q.subject, level=q.level, board=q.board, question=q.question,
            options=[q.option_a, q.option_b, q.option_c, q.option_d],
            explanation=q.explanation, correct=q.correct,
        )
        for q in rows
    ]


# ---------- admin: list all questions for a subject, grouped by level in the frontend ----------
@router.get("/admin/questions", response_model=List[schemas.QuestionOut])
def admin_list_questions(subject: Optional[str] = None, db: Session = Depends(get_db),
                          _admin: models.User = Depends(require_admin)):
    q = db.query(models.Question)
    if subject:
        q = q.filter(models.Question.subject == subject)
    return q.order_by(models.Question.subject, models.Question.level).all()


@router.post("/admin/questions", response_model=schemas.QuestionOut, status_code=status.HTTP_201_CREATED)
def admin_add_question(payload: schemas.QuestionCreate, db: Session = Depends(get_db),
                        _admin: models.User = Depends(require_admin)):
    if payload.subject not in VALID_SUBJECTS:
        raise HTTPException(status_code=400, detail="Unknown subject")
    q = models.Question(
        subject=payload.subject, level=payload.level, board=payload.board, question=payload.question,
        option_a=payload.option_a, option_b=payload.option_b, option_c=payload.option_c, option_d=payload.option_d,
        correct=payload.correct, explanation=payload.explanation,
    )
    db.add(q)
    db.commit()
    db.refresh(q)
    return q


@router.delete("/admin/questions/{question_id}", status_code=status.HTTP_204_NO_CONTENT)
def admin_delete_question(question_id: int, db: Session = Depends(get_db),
                           _admin: models.User = Depends(require_admin)):
    q = db.query(models.Question).filter(models.Question.id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    db.delete(q)
    db.commit()
    return None


# ---------- admin: bulk upload from an Excel sheet ----------
# Expected columns (case-insensitive, order doesn't matter):
#   Subject | Level (1-5) | Board | Question | OptionA | OptionB | OptionC | OptionD | Correct (A-D or 1-4) | Explanation
# The Level column is what decides which level on the student's map a question
# belongs to, and therefore the score the student must clear to pass that level.
@router.post("/admin/questions/upload", response_model=schemas.ExcelUploadResult)
def admin_upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db),
                        _admin: models.User = Depends(require_admin)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx or .xls file")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read that file — is it a valid Excel file?")

    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="The sheet is empty")

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    idx = {
        "subject": col("subject"),
        "level": col("level"),
        "board": col("board"),
        "question": col("question", "q"),
        "a": col("optiona", "option a", "a"),
        "b": col("optionb", "option b", "b"),
        "c": col("optionc", "option c", "c"),
        "d": col("optiond", "option d", "d"),
        "correct": col("correct", "answer"),
        "explanation": col("explanation", "expl"),
    }
    required = ["subject", "level", "question", "a", "b", "c", "d", "correct"]
    missing = [r for r in required if idx[r] is None]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required column(s): {', '.join(missing)}")

    added, skipped, errors = 0, 0, []
    for row_num, row in enumerate(rows[1:], start=2):
        def get(key):
            i = idx[key]
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        subject = get("subject").lower()
        level_raw = get("level")
        board = get("board") or "CBSE"
        question_text = get("question")
        a, b, c, d = get("a"), get("b"), get("c"), get("d")
        explanation = get("explanation")
        correct_raw = get("correct").upper()

        if subject not in VALID_SUBJECTS:
            skipped += 1
            errors.append(f"Row {row_num}: unknown subject '{subject}'")
            continue
        try:
            level = int(float(level_raw))
        except ValueError:
            skipped += 1
            errors.append(f"Row {row_num}: invalid level '{level_raw}'")
            continue
        if not (1 <= level <= 5):
            skipped += 1
            errors.append(f"Row {row_num}: level must be 1-5")
            continue
        if not (question_text and a and b and c and d):
            skipped += 1
            errors.append(f"Row {row_num}: missing question text or options")
            continue
        correct = LETTER_MAP.get(correct_raw)
        if correct is None:
            skipped += 1
            errors.append(f"Row {row_num}: invalid correct answer '{correct_raw}'")
            continue

        db.add(models.Question(
            subject=subject, level=level, board=board, question=question_text,
            option_a=a, option_b=b, option_c=c, option_d=d, correct=correct, explanation=explanation,
        ))
        added += 1

    db.commit()
    return schemas.ExcelUploadResult(added=added, skipped=skipped, errors=errors[:50])
